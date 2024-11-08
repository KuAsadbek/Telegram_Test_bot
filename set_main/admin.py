from django.contrib import admin
from unfold.admin import ModelAdmin
from django.http import HttpResponse
from openpyxl import Workbook
from . import models
from django.utils.html import format_html
from django.contrib import admin
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth.models import User, Group

admin.site.unregister(User)
admin.site.unregister(Group)
# Синхронный токен бота

def export_to_excel(modeladmin, request, queryset):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'User Results'

    # Стиль заголовка
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Заголовки
    headers = ['User', 'Test', 'Correct Count', 'Wrong Count', 'Percent', 'Correct Answers', 'Wrong Answers', 'Date']
    sheet.append(headers)

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    for row, result in enumerate(queryset, start=2):
        sheet.append([
            str(result.user),
            str(result.test),
            result.count_correct,
            result.count_wrong,
            result.percent,
            result.correct_str,
            result.wrong_str,
            result.date.strftime('%Y-%m-%d %H:%M:%S')
        ])
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Автоширина столбцов
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Создаем и сохраняем HTTP-ответ с файлом Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="user_results.xlsx"'
    workbook.save(response)
    return response

export_to_excel.short_description = "Excel file"

@admin.register(models.BotToken)
class Bot_Token(ModelAdmin):
    exclude = ('user',)  # Исключаем поле user из формы
    list_display = ('user','user_name','token',)  # Не показываем поле user в списке
    def has_add_permission(self, request):
        return False  # Запрещаем добавление
    def has_delete_permission(self, request, obj=None):
        return False  # Запрещаем удаление

@admin.register(models.BotMessage)
class Bot_Message(ModelAdmin):
    readonly_fields = ('command',)
    list_display = ('command', 'text', 'photo')
    def has_add_permission(self, request):
        return False  # Запрещаем добавление
    def has_delete_permission(self, request, obj=None):
        return False  # Запрещаем удаление 

@admin.register(models.CreateTest)
class create_test(ModelAdmin):
    list_display = ('cod', 'subject', 'test', 'finish_date')
    def get_exclude(self, request, obj=None):
        if obj and not obj.subject:  # Если объект существует и поле `subject` пустое
            return ['subject']  # Скрываем `subject` при редактировании
        return []

@admin.register(models.ChanelGroup)
class chanel_group(ModelAdmin):
    list_display = ('group_name', 'group_id', 'group_url')

@admin.register(models.BotButtonInline)
class Bot_Button_Inline(ModelAdmin):
    list_display = ('display_message', 'text', 'callback_data')
    readonly_fields = ('display_message', 'callback_data')
    def display_message(self, obj):
        # Форматируем строку без ссылки
        return format_html('<span>{}</span>', obj.message)
    display_message.short_description = 'Message'
    def get_readonly_fields(self, request, obj=None):
        # Делаем поля только для чтения, если объект уже существует
        if obj:  # Если объект существует (при редактировании)
            return ['display_message', 'callback_data']  # Указываем, что оба поля readonly
        return []
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('message')  # Используем join для поля 'message'
    def get_fields(self, request, obj=None):
        # Удаляем поле 'message' из отображаемых полей, оставляя только 'display_message'
        fields = super().get_fields(request, obj)
        if obj:  # Если объект существует, убираем поле 'message'
            fields.remove('message')  # Убираем поле 'message' для редактирования
        return fields

@admin.register(models.BotButtonReply)
class Bot_Reply(ModelAdmin):
    readonly_fields = ('message',)
    list_display = ('message', 'text')

    def has_add_permission(self, request):
        return False  # Запрещаем добавление

    def has_delete_permission(self, request, obj=None):
        return False  # Запрещаем удаление

@admin.register(models.UserCreate)
class User_create(ModelAdmin):
    list_display = ('telegram_id', 'first_name', 'last_name', 'schools', 'school', 'teacher_name', 'teacher_last', 'number')

@admin.register(models.UserResult)
class ResultsAdmin(ModelAdmin):
    readonly_fields = ('user', 'test',)
    list_display = ('user', 'test', 'count_correct', 'count_wrong', 'percent', 'correct_str', 'wrong_str', 'date')
    actions = [export_to_excel]
    actions_on_top = True
    actions_on_bottom = True
    exclude = ('date',)
    
    fieldsets = (
        (None, {
            'fields': ('user', 'test', 'count_correct', 'count_wrong', 'percent', 'correct_str', 'wrong_str')
        }),
    )